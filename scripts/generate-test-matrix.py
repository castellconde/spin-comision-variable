#!/usr/bin/env python3
"""
generate-test-matrix.py — Convierte la Matriz de Pruebas Unitarias en Markdown
(producida por el agente doc-writer) a un .xlsx con header rojo OXXO, casos
coloreados PASA/FALLA y hoja de Resumen con totales.

Uso:
    python scripts/generate-test-matrix.py docs/Matriz_Pruebas_Unitarias_v0.1.md
    python scripts/generate-test-matrix.py docs/Matriz_Pruebas_Unitarias_v0.1.md --output out.xlsx

Dependencias:
    pip install openpyxl
"""
import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: falta openpyxl. Instala con: pip install openpyxl")

OXXO_RED = "E2001A"
WHITE = "FFFFFF"
PASA_FILL = "C6EFCE"
PASA_FONT = "006100"
FALLA_FILL = "FFC7CE"
FALLA_FONT = "9C0006"
HEADER_GRAY = "D9D9D9"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_markdown_tables(md_text):
    """Devuelve lista de tablas; cada tabla es lista de filas (lista de celdas)."""
    tables = []
    lines = md_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.startswith("|") and i + 1 < len(lines) and re.match(r"^\|[\s\-:|]+\|$", lines[i + 1]):
            rows = [[c.strip() for c in line.strip("|").split("|")]]
            i += 2
            while i < len(lines) and lines[i].startswith("|"):
                rows.append([c.strip() for c in lines[i].strip("|").split("|")])
                i += 1
            tables.append(rows)
        else:
            i += 1
    return tables


def pick_matrix_table(tables):
    """Elige la tabla de casos: la que tenga una columna 'Caso de prueba' o 'Pasa'."""
    for t in tables:
        header = [h.lower() for h in t[0]]
        if any("caso de prueba" in h for h in header) or any("pasa" in h for h in header):
            return t
    # fallback: la tabla más grande
    return max(tables, key=len) if tables else None


def find_status_col(header):
    for idx, h in enumerate(header):
        hl = h.lower()
        if "pasa" in hl or "falla" in hl or "resultado esperado" == hl:
            return idx
    return None


def is_pasa(value):
    v = value.strip().upper()
    return v in ("PASA", "✓", "OK", "PASS") or v.startswith("PASA")


def build_xlsx(table, project_title, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz"

    header = table[0]
    data = table[1:]
    n_cols = len(header)
    status_col = find_status_col(header)
    # columna 'Pasa/Falla' explícita tiene prioridad para colorear
    pasa_falla_col = None
    for idx, h in enumerate(header):
        if "pasa" in h.lower() and "falla" in h.lower():
            pasa_falla_col = idx
    color_col = pasa_falla_col if pasa_falla_col is not None else None

    # Fila 1: título OXXO
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"MATRIZ DE PRUEBAS UNITARIAS — {project_title}")
    title_cell.fill = PatternFill("solid", fgColor=OXXO_RED)
    title_cell.font = Font(bold=True, color=WHITE, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Fila 2: encabezados de columna (rojo OXXO)
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=OXXO_RED)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Datos
    total = pasa = falla = 0
    for r, row in enumerate(data, start=3):
        for c in range(n_cols):
            val = row[c] if c < len(row) else ""
            cell = ws.cell(row=r, column=c + 1, value=val)
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
        if color_col is not None and color_col < len(row):
            total += 1
            cell = ws.cell(row=r, column=color_col + 1)
            if is_pasa(row[color_col]):
                pasa += 1
                cell.fill = PatternFill("solid", fgColor=PASA_FILL)
                cell.font = Font(size=9, bold=True, color=PASA_FONT)
            else:
                falla += 1
                cell.fill = PatternFill("solid", fgColor=FALLA_FILL)
                cell.font = Font(size=9, bold=True, color=FALLA_FONT)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Anchos de columna
    widths = []
    for c in range(n_cols):
        maxlen = max([len(str(header[c]))] + [len(str(row[c])) for row in data if c < len(row)] + [8])
        widths.append(min(max(maxlen + 2, 10), 50))
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"

    # Hoja Resumen
    rs = wb.create_sheet("Resumen")
    rs.merge_cells("A1:B1")
    t = rs.cell(row=1, column=1, value="RESUMEN DE PRUEBAS")
    t.fill = PatternFill("solid", fgColor=OXXO_RED)
    t.font = Font(bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal="center")
    resumen = [
        ("Total de casos", total),
        ("PASA", pasa),
        ("FALLA", falla),
        ("% PASA", f"{(pasa / total * 100):.1f}%" if total else "N/A"),
    ]
    for i, (k, v) in enumerate(resumen, start=2):
        kc = rs.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True)
        kc.fill = PatternFill("solid", fgColor=HEADER_GRAY)
        kc.border = BORDER
        vc = rs.cell(row=i, column=2, value=v)
        vc.border = BORDER
        if k == "PASA":
            vc.fill = PatternFill("solid", fgColor=PASA_FILL); vc.font = Font(bold=True, color=PASA_FONT)
        elif k == "FALLA":
            vc.fill = PatternFill("solid", fgColor=FALLA_FILL); vc.font = Font(bold=True, color=FALLA_FONT)
    rs.column_dimensions["A"].width = 20
    rs.column_dimensions["B"].width = 14

    wb.save(str(output_path))
    return total, pasa, falla


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("md_file", help="Markdown de la Matriz de Pruebas Unitarias")
    ap.add_argument("--output", help="Archivo .xlsx de salida (default: mismo nombre .xlsx)")
    ap.add_argument("--title", default="SPIN Comisión Variable", help="Título del proyecto")
    args = ap.parse_args()

    md_path = Path(args.md_file)
    if not md_path.exists():
        sys.exit(f"ERROR: archivo no encontrado: {md_path}")

    tables = parse_markdown_tables(md_path.read_text(encoding="utf-8"))
    table = pick_matrix_table(tables)
    if not table:
        sys.exit("ERROR: no se encontró una tabla de casos de prueba en el Markdown.")

    output_path = Path(args.output) if args.output else md_path.with_suffix(".xlsx")
    total, pasa, falla = build_xlsx(table, args.title, output_path)
    print(f"✓ Generado: {output_path}")
    print(f"  Casos: {total} | PASA: {pasa} | FALLA: {falla}")


if __name__ == "__main__":
    main()
